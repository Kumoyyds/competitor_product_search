"""Cold start workflow (spec §5.9, D22).

New site with empty parsers table:
    1. User provides an Excel sheet declaring page_type + URL
    2. Fetch pages via HTMLScraper._get_unlocker()
    3. LLM generates first-version parser from a representative HTML
    4. Sandbox-execute parser against every URL's HTML
    5. Human reviews every result and supplies structured corrections
    6. Later rounds repair and rerun all cases; the last ladder rung repeats
    7. All fetched/in-scope cases pass → seed parser + golden_samples together

Run:
    python -m src.scraping.coldstart --site tesco --input tesco.xlsx
"""

from __future__ import annotations

import argparse
import asyncio
import json
import logging
import re
import sys
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook

from .config import get_config
from .exceptions import ColdStartInputError
from .extraction import with_extraction_retry
from .models.enums import PAGE_TYPES
from .models.product_data import ProductData
from .providers import make_chat_client
from .registry import get_scrapers
from .repair.golden import _matches_expected, classify_page_type
from .repair.prepass import build_price_aware_context
from .repair.prompts import coldstart_repair_prompt, initial_parser_gen_prompt
from .repair.sandbox import run_in_sandbox
from .site_profile import (
    golden_min_for as site_golden_min_for,
    is_mandatory_page_type as site_page_type_mandatory,
    page_type_available,
)
from .storage import GoldenStore, ParserStore, ScrapeDB
from .validation import validate

logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class ColdStartRow:
    page_type: str
    url: str
    row_no: int


@dataclass(frozen=True)
class FieldCorrection:
    field: str
    correct_value: str


@dataclass(frozen=True)
class ReviewFeedback:
    url: str
    page_type: str
    corrections: tuple[FieldCorrection, ...]
    hint: str


@dataclass(frozen=True)
class _ReviewCase:
    row: ColdStartRow
    html: str
    product: Optional[ProductData]
    failure_kind: Optional[str] = None
    failure_detail: Optional[str] = None


async def run_coldstart(
    site: str,
    rows: list[ColdStartRow],
    input_fn=input,
    force_fetch: bool = False,
) -> dict[str, Any]:
    """End-to-end cold start for a site.

    Args:
        site: site identifier (e.g. "tesco")
        rows: validated page_type + URL declarations
        input_fn: injectable input function for testing (defaults to builtins.input)
        force_fetch: bypass reusable non-stale golden HTML snapshots

    Returns:
        Cold-start result including coverage_shortfall.
    """
    scraper_cls = _pick_html_scraper(site)
    scraper = scraper_cls()
    cfg = get_config()
    ladder = cfg.cold_start_model_ladder
    temperatures = cfg.cold_start_temperature_ladder
    assert ladder, "cold_start_model_ladder must contain at least one model"
    assert len(temperatures) == len(ladder), (
        f"cold_start_temperature_ladder length ({len(temperatures)}) != "
        f"cold_start_model_ladder length ({len(ladder)})"
    )

    print(f"\n== Cold start for site '{site}' ==")
    print(f"Resolving {len(rows)} URLs via golden cache / {scraper_cls.__name__}...")

    fetched = await _resolve_html(scraper, rows, force_fetch=force_fetch)
    print(
        f"Resolved {len(fetched)} pages "
        f"({sum(1 for _, status, _, _ in fetched if status == 200)} OK)."
    )

    representative_html = _pick_representative_html(fetched)
    if representative_html is None:
        print("No valid HTML fetched — aborting cold start.")
        return _empty_result(aborted=True)

    existing_coverage = _get_existing_coverage(site)
    round_feedbacks: list[ReviewFeedback] = []
    round_failures: list[str] = []
    reported: dict[str, set[str]] = {}
    resolved: dict[str, set[str]] = {}
    regressions: dict[str, set[str]] = {}
    previously_accepted: dict[str, ProductData] = {}
    parser_code: Optional[str] = None
    accepted: list[tuple[ColdStartRow, str, ProductData]] = []
    round_index = 0
    consecutive_empty = 0
    partial = False
    aborted = False

    while True:
        rung = min(round_index, len(ladder) - 1)
        model, temperature = ladder[rung], temperatures[rung]
        enable_thinking = round_index >= len(ladder) - 1
        role = "last" if enable_thinking else "middle"
        # There is nothing to repair until a node returns usable code, so a node
        # that produced nothing (truncated reply, unparsable JSON) leaves the
        # next round generating from scratch rather than discarding useful state.
        if parser_code is None:
            print(f"\nGenerating first-version parser via {model}...")
            new_code = await _gen_initial_parser(
                site,
                representative_html,
                model=model,
                temperature=temperature,
                enable_thinking=enable_thinking,
            )
        else:
            print(f"\nRepairing parser via {model} (round {round_index + 1})...")
            new_code = await _gen_repaired_parser(
                site=site,
                html=representative_html,
                current_code=parser_code,
                feedbacks=round_feedbacks,
                failures=round_failures,
                index=round_index,
                model=model,
                temperature=temperature,
                enable_thinking=enable_thinking,
                role=role,
                resolved_ledger=_sorted_ledger(resolved),
                regressions=_sorted_ledger(regressions),
            )

        if not new_code:
            consecutive_empty += 1
            if consecutive_empty >= 2:
                print("LLM returned no usable parser code twice consecutively — aborting.")
                return _failed_result(existing_coverage, site=site, aborted=True)
            print(
                f"{model} returned no usable parser code — "
                "falling through to the next repair round."
            )
            round_index += 1
            continue
        consecutive_empty = 0
        parser_code = new_code

        print("\nRunning parser against all resolved HTMLs (sandbox)...")
        cases = await _run_review_cases(site, rows, fetched, parser_code)
        print(f"\n{len(cases)} cold-start cases for review.\n")

        accepted = []
        accepted_by_type: Counter[str] = Counter()
        sandbox_failed: list[tuple[ColdStartRow, str]] = []
        human_rejected: list[tuple[ColdStartRow, ReviewFeedback]] = []
        extraction_failed: list[tuple[ColdStartRow, str]] = []
        unavailable_accepted: list[tuple[str, str]] = []
        review_aborted = False

        for i, case in enumerate(cases, 1):
            row = case.row
            current_count = (
                existing_coverage.get(row.page_type, 0)
                + accepted_by_type[row.page_type]
            )
            if current_count >= cfg.golden_max_for(row.page_type):
                print(
                    f"\n[{i}/{len(cases)}] {row.url}\n"
                    f"  bucket '{row.page_type}' already has {current_count} goldens — "
                    "skipping (spare)"
                )
                continue

            print(f"\n[{i}/{len(cases)}] {row.url}")
            print(f"  declared page_type = {row.page_type}")
            if case.product is None:
                detail = case.failure_detail or "unknown failure"
                if case.failure_kind == "extraction_failed":
                    extraction_failed.append((row, detail))
                    print(f"  FETCH ERROR: {detail} — not counted against parser")
                else:
                    sandbox_failed.append((row, detail))
                    print(f"  PARSER ERROR: {detail}")
                continue

            product = case.product
            classified = classify_page_type(product)
            if classified != row.page_type:
                print(
                    f"  !! MISMATCH: extracted fields look like '{classified}', "
                    f"not declared '{row.page_type}'"
                )
            _print_product_panel(product, row.page_type)

            previous = previously_accepted.get(row.url)
            if previous is not None and _products_match(product, previous):
                accepted.append((row, case.html, product))
                accepted_by_type[row.page_type] += 1
                print("  → unchanged since prior acceptance; accepted automatically")
                if not page_type_available(site, classified):
                    unavailable_accepted.append((row.url, classified))
                continue

            ans = input_fn("  Accept? [y/N/q] ").strip().lower()
            if ans == "q":
                print("  Aborting review loop.")
                review_aborted = True
                break
            if ans == "y":
                accepted.append((row, case.html, product))
                accepted_by_type[row.page_type] += 1
                print("  → accepted")
                if not page_type_available(site, classified):
                    unavailable_accepted.append((row.url, classified))
                continue

            feedback = _collect_feedback(row, input_fn)
            human_rejected.append((row, feedback))
            print("  → rejected")

        if unavailable_accepted:
            print(
                "\n!! SITE PROFILE WARNING: accepted result(s) classified into "
                "a page type declared unavailable:"
            )
            for url, classified in unavailable_accepted:
                print(f"  - {url}: classified as '{classified}'")

        if review_aborted:
            _print_failure_summary(
                sandbox_failed, human_rejected, extraction_failed, regressions={}
            )
            return _failed_result(
                existing_coverage, site=site,
                aborted=True,
                accepted=len(accepted),
            )

        if not sandbox_failed and not human_rejected:
            if extraction_failed:
                _print_failure_summary([], [], extraction_failed, regressions={})
            parser_id, seeded_count = _seed(site, parser_code, accepted)
            break

        round_feedbacks = [feedback for _, feedback in human_rejected]
        round_failures = [
            f"{row.url} ({row.page_type}): {detail}"
            for row, detail in sandbox_failed
        ]
        regressions = _update_review_ledger(
            reported=reported,
            resolved=resolved,
            previously_accepted=set(previously_accepted),
            accepted=accepted,
            sandbox_failed=sandbox_failed,
            human_rejected=human_rejected,
        )
        _print_failure_summary(
            sandbox_failed,
            human_rejected,
            extraction_failed,
            regressions=regressions,
        )
        previously_accepted = {row.url: product for row, _, product in accepted}

        at_cap = round_index + 1 >= cfg.cold_start_max_repair_rounds
        if at_cap:
            print(
                f"\nCold-start safety cap reached "
                f"({cfg.cold_start_max_repair_rounds} rounds)."
            )
            action = input_fn(
                "Save current results or quit? "
                "[s=保存当前结果并退出 / q=放弃退出] "
            ).strip().lower()
        else:
            action = input_fn(
                "\nContinue 纠错? "
                "[c=继续修复 / s=保存当前结果并退出 / q=放弃退出] "
            ).strip().lower()

        if action == "s":
            parser_id, seeded_count = _seed(site, parser_code, accepted)
            partial = True
            break
        if not at_cap and action in {"c", "y"}:
            round_index += 1
            continue

        if at_cap:
            print("Current parser was not saved after reaching the safety cap.")
        else:
            print("Parser was not saved.")
        return _failed_result(
            existing_coverage, site=site,
            aborted=True,
            accepted=len(accepted),
        )

    final_coverage = dict(existing_coverage)
    for page_type, count in Counter(row.page_type for row, _, _ in accepted).items():
        final_coverage[page_type] = final_coverage.get(page_type, 0) + count
    coverage_shortfall = _coverage_shortfall(final_coverage, site)
    if coverage_shortfall:
        print("\nWARNING: mandatory golden coverage is still incomplete:")
        for page_type in coverage_shortfall:
            print(
                f"  - {page_type}: needs >= {site_golden_min_for(site, page_type)}, "
                f"has {final_coverage.get(page_type, 0)}"
            )
    heading = "Cold start partial result saved" if partial else "Cold start complete"
    print(f"\n== {heading} ==")
    print(f"  parser_id     = {parser_id}")
    print(f"  goldens seeded = {seeded_count}")
    return {
        "parser_id": parser_id,
        "seeded_goldens": seeded_count,
        "accepted": len(accepted),
        "aborted": aborted,
        "partial": partial,
        "coverage_shortfall": coverage_shortfall,
    }


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def _pick_html_scraper(site: str):
    from .scrapers.html_scraper import HTMLScraper

    for cls in get_scrapers(site):
        if issubclass(cls, HTMLScraper):
            return cls
    raise ValueError(f"No HTMLScraper registered for site: {site}")


def _empty_result(aborted: bool) -> dict[str, Any]:
    return {
        "parser_id": None,
        "seeded_goldens": 0,
        "accepted": 0,
        "aborted": aborted,
        "partial": False,
        "coverage_shortfall": [],
    }


def _failed_result(
    existing_coverage: dict[str, int],
    *,
    site: str = "",
    aborted: bool,
    accepted: int = 0,
) -> dict[str, Any]:
    return {
        **_empty_result(aborted=aborted),
        "accepted": accepted,
        "coverage_shortfall": _coverage_shortfall(existing_coverage, site),
    }


def _get_existing_coverage(site: str) -> dict[str, int]:
    cfg = get_config()
    db = ScrapeDB(cfg.db_path)
    db.init_db()
    try:
        return GoldenStore(db).get_page_type_coverage(site)
    finally:
        db.close()


def _coverage_shortfall(
    coverage: dict[str, int], site: str = ""
) -> list[str]:
    return [
        page_type
        for page_type in PAGE_TYPES
        if site_page_type_mandatory(site, page_type)
        and coverage.get(page_type, 0) < site_golden_min_for(site, page_type)
    ]


async def _batch_fetch(scraper, urls: list[str]) -> list[tuple[str, int, str]]:
    """Fetch all URLs; returns list of (url, status_code, html)."""
    if not urls:
        return []
    cfg = get_config()
    unlocker = scraper._get_unlocker()
    semaphore = asyncio.Semaphore(min(cfg.per_site_concurrency, len(urls)))

    async def _one(url: str) -> tuple[str, int, str]:
        async with semaphore:
            try:
                status, html = await with_extraction_retry(unlocker.fetch, url)
                return (url, status, html)
            except Exception as e:
                logger.warning("cold start fetch failed for %s: %s", url, e)
                return (url, 0, "")

    results = await asyncio.gather(*[_one(u) for u in urls])
    return list(results)


async def _resolve_html(
    scraper,
    rows: list[ColdStartRow],
    force_fetch: bool,
) -> list[tuple[str, int, str, str]]:
    """Resolve HTML from non-stale goldens first, then fetch cache misses."""
    cached: dict[str, str] = {}
    if not force_fetch:
        cfg = get_config()
        db = ScrapeDB(cfg.db_path)
        db.init_db()
        try:
            gs = GoldenStore(db)
            samples = [
                sample
                for page_type in PAGE_TYPES
                for sample in gs.get_by_site_and_type(
                    scraper.site, page_type, exclude_stale=True
                )
            ]
            for sample in sorted(samples, key=lambda item: item["id"], reverse=True):
                url = sample["expected_output"].get("url")
                html = sample.get("html_snapshot")
                if url and html and url not in cached:
                    cached[url] = html
        finally:
            db.close()

    missing_urls = list(dict.fromkeys(row.url for row in rows if row.url not in cached))
    fetched = await _batch_fetch(scraper, missing_urls) if missing_urls else []
    fetched_by_url = {url: (status, html) for url, status, html in fetched}

    resolved: list[tuple[str, int, str, str]] = []
    for row in rows:
        if row.url in cached:
            item = (row.url, 200, cached[row.url], "goldset")
        else:
            status, html = fetched_by_url.get(row.url, (0, ""))
            item = (row.url, status, html, "brightdata")
        resolved.append(item)
        print(f"  [{item[3]}] {row.url}")
    return resolved


def _pick_representative_html(
    fetched: list[tuple[Any, ...]]
) -> Optional[str]:
    """Pick the largest 200 HTML as representative for LLM parser generation."""
    ok = [
        (item[0], item[2])
        for item in fetched
        if item[1] == 200 and item[2]
    ]
    if not ok:
        return None
    return max(ok, key=lambda t: len(t[1]))[1]


async def _gen_initial_parser(
    site: str,
    html: str,
    *,
    model: Optional[str] = None,
    temperature: Optional[float] = None,
    enable_thinking: Optional[bool] = None,
) -> Optional[str]:
    """Generate a parser from scratch.

    The ladder driver passes the node it is currently on; the defaults keep the
    node-0 behaviour for direct callers.
    """
    cfg = get_config()
    ladder = cfg.cold_start_model_ladder
    temperatures = cfg.cold_start_temperature_ladder
    assert ladder, "cold_start_model_ladder must contain at least one model"
    assert len(temperatures) == len(ladder), (
        f"cold_start_temperature_ladder length ({len(temperatures)}) != "
        f"cold_start_model_ladder length ({len(ladder)})"
    )
    llm = make_chat_client(
        model=model if model is not None else ladder[0],
        temperature=temperature if temperature is not None else temperatures[0],
        enable_thinking=(
            enable_thinking if enable_thinking is not None else len(ladder) == 1
        ),
        purpose="cold start",
    )
    if llm is None:
        print("LLM provider key not set — cannot generate parser.")
        return None
    try:
        price_ctx = build_price_aware_context(html, f"https://dummy/{site}")
        resp = await llm.ainvoke(initial_parser_gen_prompt(price_ctx, site))
    except Exception as e:
        _log_llm_failure("initial parser gen", e)
        return None

    content = resp.content if hasattr(resp, "content") else str(resp)
    return _extract_parser_code(content)


async def _gen_repaired_parser(
    *,
    site: str,
    html: str,
    current_code: str,
    feedbacks: list[ReviewFeedback],
    failures: list[str],
    index: int,
    model: str,
    temperature: float,
    enable_thinking: bool,
    role: str,
    resolved_ledger: dict[str, list[str]],
    regressions: dict[str, list[str]],
) -> Optional[str]:
    llm = make_chat_client(
        model=model,
        temperature=temperature,
        enable_thinking=enable_thinking,
        purpose="cold start repair",
    )
    if llm is None:
        print("LLM provider key not set — cannot repair parser.")
        return None
    try:
        price_ctx = build_price_aware_context(html, f"https://dummy/{site}")
        prompt = coldstart_repair_prompt(
            price_ctx,
            site,
            current_code,
            feedbacks,
            failures,
            role=role,
            attempt_index=index,
            model=model,
            resolved_ledger=resolved_ledger,
            regressions=regressions,
        )
        resp = await llm.ainvoke(prompt)
    except Exception as exc:
        _log_llm_failure("cold start parser repair", exc)
        return None
    content = resp.content if hasattr(resp, "content") else str(resp)
    return _extract_parser_code(content)


def _log_llm_failure(stage: str, exc: Exception) -> None:
    """Log an LLM call failure, calling out output truncation explicitly.

    A truncated reply arrives as ``LengthFinishReasonError``: the provider hit
    the output cap mid-JSON and the OpenAI SDK discards the partial content, so
    there is nothing to salvage — the ladder must move on to the next node.
    """
    if type(exc).__name__ == "LengthFinishReasonError":
        logger.error(
            "%s hit the model output limit — the reply was truncated before the "
            "JSON closed. Raise the provider's max_output_tokens in providers.py "
            "if this repeats.",
            stage,
        )
        print(f"{stage} exceeded the model output limit (truncated reply).")
        return
    logger.exception("%s failed: %s", stage, exc)


def _extract_parser_code(content: str) -> Optional[str]:
    try:
        parsed = json.loads(content)
    except json.JSONDecodeError:
        try:
            start = content.index("{")
            end = content.rindex("}") + 1
            parsed = json.loads(content[start:end])
        except Exception:
            return None
    return parsed.get("parser_code")


async def _run_review_cases(
    site: str,
    rows: list[ColdStartRow],
    fetched: list[tuple[str, int, str, str]],
    parser_code: str,
) -> list[_ReviewCase]:
    cases: list[_ReviewCase] = []
    for row, (url, status, html, _) in zip(rows, fetched):
        if status != 200 or not html:
            cases.append(
                _ReviewCase(
                    row=row,
                    html=html,
                    product=None,
                    failure_kind="extraction_failed",
                    failure_detail=f"HTTP/status {status}",
                )
            )
            continue

        result = await run_in_sandbox(parser_code, html, url)
        if not isinstance(result, dict):
            cases.append(
                _ReviewCase(
                    row=row,
                    html=html,
                    product=None,
                    failure_kind="sandbox_failed",
                    failure_detail=_describe_sandbox_failure(result),
                )
            )
            continue

        product, errors = validate(_wrap(result, site, url))
        if product is None:
            cases.append(
                _ReviewCase(
                    row=row,
                    html=html,
                    product=None,
                    failure_kind="sandbox_failed",
                    failure_detail="gate: " + "; ".join(errors),
                )
            )
            continue
        cases.append(_ReviewCase(row=row, html=html, product=product))
    return cases


def _describe_sandbox_failure(result: Any) -> str:
    if hasattr(result, "type_name"):
        message = getattr(result, "message", "")
        return f"{result.type_name}: {message}".rstrip(": ")
    if hasattr(result, "reason"):
        return f"{type(result).__name__}: {result.reason}"
    if hasattr(result, "timeout"):
        return f"{type(result).__name__}: {result.timeout}s"
    return type(result).__name__


_TRACING_FIELDS = frozenset(
    {"url", "website", "scraped_at", "source_type", "parser_version", "raw"}
)

_PAGE_TYPE_KEY_FIELDS: dict[str, tuple[str, ...]] = {
    "standard": ("price",),
    "membership": ("membership_price",),
    "discounted": ("list_price", "price"),
    "out_of_stock": ("in_stock", "image_urls"),
    "multipack": ("variant",),
}


def _display_fields() -> list[str]:
    """Return reviewable ProductData fields in model declaration order."""
    return [
        name for name in ProductData.model_fields
        if name not in _TRACING_FIELDS
    ]


def _print_product_panel(product: ProductData, page_type: str) -> None:
    print("  Extracted:")
    key_fields = set(_PAGE_TYPE_KEY_FIELDS.get(page_type, ()))
    for field in _display_fields():
        value = getattr(product, field)
        rendered = _format_display_value(value)
        warning = ""
        if field in key_fields and value in (None, "", [], {}):
            warning = "  !! 该桶关键字段为空"
        print(f"    {field:18s} = {rendered}{warning}")


def _format_display_value(value: Any) -> str:
    if value is None:
        return "—"
    if isinstance(value, list):
        if not value:
            return "0 项"
        return f"{len(value)} 项 ({_truncate(str(value[0]), 80)})"
    if isinstance(value, str):
        return _truncate(value, 80)
    return str(value)


def _truncate(value: str, limit: int) -> str:
    return value if len(value) <= limit else value[:limit] + "…"


def _collect_feedback(row: ColdStartRow, input_fn) -> ReviewFeedback:
    fields = _display_fields()
    print("  Which fields are wrong?")
    print("    " + "  ".join(f"{i}.{field}" for i, field in enumerate(fields, 1)))
    raw = input_fn("  (comma-separated, Enter to skip) > ").strip()
    selected: list[str] = []
    for token in (part.strip() for part in raw.split(",")):
        if not token:
            continue
        if token.isdigit() and 1 <= int(token) <= len(fields):
            field = fields[int(token) - 1]
        elif token in fields:
            field = token
        else:
            print(f"  note: ignored unknown field selection {token!r}")
            continue
        if field not in selected:
            selected.append(field)

    corrections = tuple(
        FieldCorrection(
            field=field,
            correct_value=_normalize_correction_value(
                input_fn(f"  correct {field}? > ").strip()
            ),
        )
        for field in selected
    )
    hint = input_fn("  Any hint? (Enter to skip) > ").strip()
    return ReviewFeedback(
        url=row.url,
        page_type=row.page_type,
        corrections=corrections,
        hint=hint,
    )


def _normalize_correction_value(value: str) -> str:
    """Render common human 'clear this field' input unambiguously for the LLM."""
    if value.lower() in {"-", "none", "null", "n/a", "na"}:
        return "None (clear or omit this field)"
    return value


def _products_match(current: ProductData, previous: ProductData) -> bool:
    return _matches_expected(
        current.model_dump(mode="json"), previous.model_dump(mode="json")
    ) is None


def _sorted_ledger(ledger: dict[str, set[str]]) -> dict[str, list[str]]:
    """Return stable, prompt-ready ledger entries and omit empty URL rows."""
    return {
        url: sorted(fields)
        for url, fields in sorted(ledger.items())
        if fields
    }


def _update_review_ledger(
    *,
    reported: dict[str, set[str]],
    resolved: dict[str, set[str]],
    previously_accepted: set[str],
    accepted: list[tuple[ColdStartRow, str, ProductData]],
    sandbox_failed: list[tuple[ColdStartRow, str]],
    human_rejected: list[tuple[ColdStartRow, ReviewFeedback]],
) -> dict[str, set[str]]:
    """Apply one review round to the compact fixed/regressed field ledger."""
    current_regressions: dict[str, set[str]] = {}
    wrong_by_url: dict[str, set[str]] = {}
    for row, feedback in human_rejected:
        wrong_by_url.setdefault(row.url, set()).update(
            correction.field for correction in feedback.corrections
        )
    for row, _ in sandbox_failed:
        wrong_by_url.setdefault(row.url, set()).add("<sandbox>")

    for url, wrong_fields in wrong_by_url.items():
        regressed = resolved.get(url, set()).intersection(wrong_fields)
        if url in previously_accepted:
            regressed.update(wrong_fields)
            if "<sandbox>" in wrong_fields:
                regressed.update(resolved.get(url, set()))
        if regressed:
            current_regressions[url] = set(regressed)
            resolved.setdefault(url, set()).difference_update(regressed)
        reported.setdefault(url, set()).update(wrong_fields)

    for row, _, _ in accepted:
        prior_reports = reported.get(row.url)
        if prior_reports:
            resolved.setdefault(row.url, set()).update(prior_reports)

    return current_regressions


def _print_failure_summary(
    sandbox_failed: list[tuple[ColdStartRow, str]],
    human_rejected: list[tuple[ColdStartRow, ReviewFeedback]],
    extraction_failed: list[tuple[ColdStartRow, str]],
    *,
    regressions: dict[str, set[str]],
) -> None:
    failed_count = len(sandbox_failed) + len(human_rejected)
    if failed_count:
        print(f"\nParser has {failed_count} cold-start case(s) that did not pass:")
        for row, detail in sandbox_failed:
            print(f"  [PARSER CRASH] {row.url} ({row.page_type})  {detail}")
        for row, feedback in human_rejected:
            wrong = ", ".join(c.field for c in feedback.corrections) or "unspecified"
            print(f"  [REJECTED]     {row.url} ({row.page_type})  wrong: {wrong}")
        for url, fields in sorted(regressions.items()):
            print(f"  [REGRESSION]   {url}  regressed: {', '.join(sorted(fields))}")
    if extraction_failed:
        print("\nThese URLs could not be fetched (not the parser's fault):")
        for row, detail in extraction_failed:
            print(f"  [FETCH FAIL]   {row.url} ({row.page_type})  {detail}")


def _wrap(raw: dict[str, Any], site: str, url: str) -> dict[str, Any]:
    wrapped = dict(raw)
    wrapped.setdefault("url", url)
    wrapped["website"] = site
    wrapped["source_type"] = "html"
    wrapped["scraped_at"] = datetime.now(timezone.utc)
    wrapped["parser_version"] = "coldstart_v1"
    return wrapped


def _seed(
    site: str,
    parser_code: str,
    accepted: list[tuple[ColdStartRow, str, ProductData]],
) -> tuple[int, int]:
    cfg = get_config()
    db = ScrapeDB(cfg.db_path)
    db.init_db()
    try:
        version = f"cs_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}"
        parser_id = ParserStore(db).create(
            site=site, version=version, code=parser_code, created_by="initial"
        )

        gs = GoldenStore(db)
        seeded = 0
        for row, html, product in accepted:
            gs.seed(
                site,
                row.page_type,
                html,
                product.model_dump(mode="json"),
                created_by="coldstart",
            )
            seeded += 1

        return parser_id, seeded
    finally:
        db.close()


# ---------------------------------------------------------------------------
# CLI entrypoint
# ---------------------------------------------------------------------------

def read_coldstart_input(path: Path, site: str = "") -> list[ColdStartRow]:
    """Read and validate a cold-start workbook against site-aware policy.

    The empty-site default preserves the pre-M23 global-policy behavior for
    callers that do not yet have a site identifier.
    """
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ColdStartInputError(
            f"expected an Excel file (.xlsx/.xlsm), got '{path.name}'\n"
            "  cold start input is a sheet with columns: page_type, url"
        )

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ColdStartInputError(f"could not read Excel file '{path}': {exc}") from exc

    try:
        sheet = workbook.active
        header_values = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        normalized_headers = [
            str(value).strip().lower() if value is not None else ""
            for value in header_values
        ]
        header_map = {
            header: index
            for index, header in enumerate(normalized_headers)
            if header and header not in normalized_headers[:index]
        }
        missing = [name for name in ("page_type", "url") if name not in header_map]
        if missing:
            found = (
                ", ".join(header for header in normalized_headers if header)
                or "(empty)"
            )
            raise ColdStartInputError(
                f"missing required column(s): {', '.join(missing)}\n"
                f"  found header: {found}\n"
                "  the sheet's first row must contain: page_type, url"
            )

        parsed: list[ColdStartRow] = []
        problems: list[str] = []
        for row_no, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
            page_index = header_map["page_type"]
            url_index = header_map["url"]
            page_value = values[page_index] if page_index < len(values) else None
            url_value = values[url_index] if url_index < len(values) else None
            raw_page_type = "" if page_value is None else str(page_value).strip()
            url = "" if url_value is None else str(url_value).strip()
            if not raw_page_type and not url:
                continue

            page_type = re.sub(r"[\s-]+", "_", raw_page_type.lower())
            row_problems = []
            if page_type not in PAGE_TYPES:
                row_problems.append(f"page_type={raw_page_type!r}")
            elif not page_type_available(site, page_type):
                row_problems.append(
                    f"page_type={page_type!r} is unavailable for site {site!r}"
                )
            if not url:
                row_problems.append("url='' (empty)")
            if row_problems:
                problems.append(f"  row {row_no}: " + ", ".join(row_problems))
                continue
            parsed.append(ColdStartRow(page_type=page_type, url=url, row_no=row_no))

        if problems:
            raise ColdStartInputError(
                f"invalid row value(s) in {path.name}:\n"
                + "\n".join(problems)
                + f"\nLegal page_type values: {', '.join(PAGE_TYPES)}"
            )
    finally:
        workbook.close()

    cfg = get_config()
    counts = Counter(row.page_type for row in parsed)
    mandatory = tuple(
        pt for pt in PAGE_TYPES if site_page_type_mandatory(site, pt)
    )
    missing_types = [pt for pt in mandatory if counts[pt] == 0]
    if missing_types:
        optional = [pt for pt in PAGE_TYPES if not site_page_type_mandatory(site, pt)]
        found = (
            ", ".join(f"{pt} x{counts[pt]}" for pt in PAGE_TYPES if counts[pt])
            or "none"
        )
        raise ColdStartInputError(
            f"{path.name} is missing required page_type(s): {', '.join(missing_types)}\n"
            "  cold start needs >=1 URL for each of: "
            f"{', '.join(mandatory)}\n"
            f"  optional (may be omitted): {', '.join(optional) if optional else 'none'}\n"
            f"  found: {found}\n"
            "  (edit sites.yaml for a declared site, or the global fallback "
            "in config.py)"
        )

    for page_type in PAGE_TYPES:
        maximum = cfg.golden_max_for(page_type)
        if counts[page_type] > maximum:
            print(
                f"note: {counts[page_type]} rows for '{page_type}'; at most {maximum} "
                "goldens are seeded per page_type — extras are spares\n"
                "      (used only if an earlier one fails extraction or is rejected during review)"
            )
    return parsed


def _result_exit_code(result: dict[str, Any]) -> int:
    if result.get("partial"):
        return 2
    if result.get("aborted") or result.get("parser_id") is None:
        return 1
    if result.get("coverage_shortfall"):
        return 2
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description="Cold start for a new site.")
    parser.add_argument("--site", required=True, help="site identifier (e.g. tesco)")
    input_group = parser.add_mutually_exclusive_group(required=True)
    input_group.add_argument(
        "--input",
        dest="input_path",
        type=Path,
        help="Excel workbook with page_type and url columns",
    )
    input_group.add_argument(
        "--urls-file",
        dest="input_path",
        type=Path,
        help=argparse.SUPPRESS,
    )
    parser.add_argument("--verbose", "-v", action="store_true")
    parser.add_argument(
        "--force-fetch",
        action="store_true",
        help="ignore reusable golden HTML snapshots and fetch every URL",
    )
    args = parser.parse_args()

    if args.verbose:
        logging.basicConfig(level=logging.INFO)

    try:
        rows = read_coldstart_input(args.input_path, args.site)
    except ColdStartInputError as exc:
        print(f"ColdStartInputError: {exc}")
        return 1
    if not rows:
        print(f"No data rows in {args.input_path}")
        return 1

    result = asyncio.run(
        run_coldstart(args.site, rows, force_fetch=args.force_fetch)
    )
    return _result_exit_code(result)


if __name__ == "__main__":
    sys.exit(main())
